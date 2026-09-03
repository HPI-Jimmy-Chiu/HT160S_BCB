#!/usr/bin/env python
# -*- coding: utf-8 -*-
r"""gen-alid-catalog.py -- the ONE generator for every HT160S ALID artifact.

Home in the repo:  scripts/ops/gen-alid-catalog.py   (python 3.8+; checked in 2026-09-03, S1)
Sibling precedent in that folder: check-ladder-consistency.py, md-manual-to-docx.py.

WHY THIS EXISTS
    Before this script the 481 ALID rows in the customer workbook and the 481
    ALID_CATALOG tuples in the SECS host simulator were HAND-PASTED, so they
    drifted: workbook and simulator sat at 481 rows while the machine had 485.
    Everything downstream of system/AlarmList.csv is now generated.

        system/AlarmList.csv   (startup dump of SYSTEM_MODULAR::mapAlarmCodeList)
              |
              +--> docs/SECS/SECS_GEM功能_Handler_<date>.xlsx   sheet "ALID"
              +--> <sim>/code/secs_host_simulator.py            ALID_CATALOG
              +--> <sim>/code/scenario_runner.py                ALARM_CATALOG_ROWS
              +--> docs/SECS/HT160S_ALID_map_<date>.csv         review / delivery

ALID RULE -- "Option D", class-banded 9-digit ALID.
    Owner-ratified 2026-09-02: docs/plan/alid-option-d-ratified-spec-20260902.md
    THIS FILE MUST STAY IN LOCK-STEP WITH ComputeAlarmAlid() in
    HT160S_Program_BCB_V1.0.0.0/SecsGem/UsecegemMainFrom.cpp:167-177.

        ALID = Class * 100,000,000 + Payload          # ALWAYS exactly 9 digits
        class 1/2/3 : code prefix JAM / WAR / MES, payload = the numeric tail
        class 4..8  : EXACTLY-5-char numeric code, first char 4..8,
                      payload = the whole 5-digit code
                      4=cylinder 5=motor 6=sucker 7=eRecordProcess 8=eOther
        class 9     : anything else. payload = the legacy 31-poly hash folded to
                      8 digits. Class 9 MEANS "NOT in the S5F6/S5F8 catalog --
                      read the alarm code from the leading token of ALTX".
        class 0     : never produced (it would be shorter than 9 digits).

    AMENDMENT 1   classes 1/2/3 reject payload > 999,999 (fall to class 9).
                  Under Amendment 1b this can never fire; kept as a guard.
    AMENDMENT 1b  classes 1/2/3 require a CANONICAL tail -- exactly one of
                      len(tail) == 4 and value <  10000     ("0913", "1421")
                      len(tail) == 5 and value >= 10000     ("16120")
                  Anything else -> class 9. This pins one payload value to one
                  canonical string (pre-1b MES1421 / MES01421 / MES001421 all
                  composed 300001421). It is NOT "reject any leading zero" --
                  that naive rule kills 11 in-service unit-01..09 codes.
                  Consequence: units 01-09 are capped at 100 alarms each
                  ("0UNN"); "0UNNN" is illegal, so WAR09120 is NOT allowed and
                  such a unit must be renumbered >= 10.
    AMENDMENT A5  classes 7 and 8 are LIVE numeric classes (was: reserved).
                  database.h:23 eRecordProcess=7, database.h:24 eOther=8.
    AMENDMENT 2   startup self-check in CreateSystemAlarmCode(); --audit below
                  is its offline twin (same three rules).

DATA-SOURCE GATE (B1, added 2026-09-03)
    system/AlarmList.csv is a BOOT ARTIFACT (SYSTEM_MODULAR::CreateSystemAlarmCode
    dumps mapAlarmCodeList at every start), yet it is what everything below is
    derived from. The committed copy can lag the source: on 2026-09-02 HEAD held
    484 rows (WAR0963 missing) while a machine that had booted the same source
    wrote 485. Regenerating from a clean checkout that was never booted would
    have shipped a 484-row dictionary to the customer.
    Therefore EVERY run first asserts, against the constants right below:
        EXPECTED_CSV_ROWS  -- the exact number of alarm codes
        EXPECTED_CSV_MD5   -- the md5 of the whole file
    A mismatch is fatal. When database.cpp legitimately registers a new code:
    boot the sim once so AlarmList.csv is rewritten, then bump BOTH constants in
    the SAME commit as the database.cpp change. --expect-csv-md5 / --expect-csv-rows
    override the constants for that one run; --allow-csv-drift downgrades the
    gate to a warning (for exploration only -- never for an emit that is shipped).

USAGE
    python scripts/ops/gen-alid-catalog.py --audit
    python scripts/ops/gen-alid-catalog.py --selftest
    python scripts/ops/gen-alid-catalog.py --emit-sheet-csv out\ALID_sheet.csv
    python scripts/ops/gen-alid-catalog.py --emit-map-csv  docs\SECS\HT160S_ALID_map_20260902.csv
    python scripts/ops/gen-alid-catalog.py --emit-sim-block out\ALID_CATALOG.py
    python scripts/ops/gen-alid-catalog.py --emit-sim --in-place
    python scripts/ops/gen-alid-catalog.py --emit-xlsx "docs\SECS\SECS_GEM功能_Handler_20260902.xlsx"
        --xlsx-base "docs\SECS\SECS_GEM功能_Handler_20260831.xlsx"
        --expect-base-md5 e0ff9e2c182dc6521fd706a2db00abda
    python scripts/ops/gen-alid-catalog.py --check
        CI mode: regenerate in memory, diff against what is on disk. Exit 1 if stale.
    python scripts/ops/gen-alid-catalog.py --verify-table <amended_mapping.csv>
        Cross-check every row against an externally produced old->new table.
"""
import argparse
import csv
import hashlib
import io
import os
import re
import sys
from collections import Counter

# --------------------------------------------------------------------------- #
# Option D core -- the reference implementation the BCB6 port must mirror.
# --------------------------------------------------------------------------- #
M32 = 1 << 32
BAND = 100000000
# B1 data-source gate -- bump both in the SAME commit that registers a new alarm
# code in database.cpp (after booting the sim once so AlarmList.csv is rewritten).
EXPECTED_CSV_ROWS = 486   # 485 + MES0926 (owner ruling D2, 2026-09-03)
EXPECTED_CSV_MD5 = "b710227ca23356f178c7384bf2784e70"   # system/AlarmList.csv after the 2026-09-03 sim boot
DIG = bytearray(b"0123456789")
UPP = bytearray(b"ABCDEFGHIJKLMNOPQRSTUVWXYZ")
PREFIX_CLASS = {b"JAM": 1, b"WAR": 2, b"MES": 3}
NUM_CLASS = {ord("4"): 4, ord("5"): 5, ord("6"): 6,
             ord("7"): 7, ord("8"): 8}       # AMENDMENT A5 widened 4/5/6 -> 4..8
PREFIX_OF = {1: "JAM", 2: "WAR", 3: "MES"}
CLASS_NAME = {0: "never produced", 1: "JAM", 2: "WAR", 3: "MES",
              4: "cylinder 4xxxx", 5: "motor 5xxxx", 6: "sucker 6xxxx",
              7: "eRecordProcess 7xxxx", 8: "eOther 8xxxx", 9: "free string"}
CLASS_SLOTS = {1: 100000, 2: 100000, 3: 100000, 4: 10000, 5: 10000,
               6: 10000, 7: 10000, 8: 10000, 9: 100000000}


def legacy_hash(b):
    """The pre-Option-D producer: unsigned 32-bit 31-polynomial over code bytes."""
    a = 0
    for ch in bytearray(b):
        a = (a * 31 + ch) % M32
    return a


def canonical_tail(tail):
    """AMENDMENT 1b. Exactly one of len==4 & v<10000 | len==5 & v>=10000."""
    n = len(tail)
    if n == 4 or n == 5:
        v = int(tail)
        if n == 4 and v < 10000:
            return True, v
        if n == 5 and v >= 10000:
            return True, v
        return False, v
    return False, -1


def classify(b):
    """b = raw code bytes (AnsiString semantics). -> (cls, payload, reason)."""
    bb = bytearray(b)
    n = len(bb)
    # shape A : exactly 5 chars, all digits, first char 4..8
    if n == 5 and all(c in DIG for c in bb) and bb[0] in NUM_CLASS:
        return NUM_CLASS[bb[0]], int(bytes(bb)), "numeric5"
    # shape B : 3 uppercase letters + an all-digit non-empty tail
    if (n >= 4 and bb[0] in UPP and bb[1] in UPP and bb[2] in UPP
            and all(c in DIG for c in bb[3:])):
        cls = PREFIX_CLASS.get(bytes(bb[0:3]))
        if cls is None:
            return 9, legacy_hash(b) % BAND, "unknown_prefix"
        ok, payload = canonical_tail(bytes(bb[3:]))
        if not ok:                                       # AMENDMENT 1b
            return 9, legacy_hash(b) % BAND, "amend1b_noncanonical"
        if payload > 999999:                             # AMENDMENT 1 (unreachable)
            return 9, legacy_hash(b) % BAND, "amend1_overflow"
        return cls, payload, "prefixed"
    return 9, legacy_hash(b) % BAND, "freestring"


def optd(b):
    cls, payload, _ = classify(b)
    return cls * BAND + payload


def decode(alid):
    """The KYEC EAP side. -> (class, payload, canonical_code or None)."""
    cls, pay = alid // BAND, alid % BAND
    if cls in PREFIX_OF:                     # 1b makes this total and injective
        return cls, pay, "%s%0*d" % (PREFIX_OF[cls], 4 if pay < 10000 else 5, pay)
    if cls in (4, 5, 6, 7, 8):
        return cls, pay, "%05d" % pay
    return cls, pay, None


# --------------------------------------------------------------------------- #
# the alarm universe
# --------------------------------------------------------------------------- #
CSV_COLS = ["AlarmCode", "AlarmType", "E_ErrMessage", "C_ErrMessage",
            "E_Description", "C_Description"]
# The 類別 label the workbook shows per ALCD, verbatim from the 0831 ALID sheet.
ALCD_LABEL = {0: "Jam 卡料", 1: "Message 訊息", 4: "Cylinder 汽缸",
              5: "Motor 馬達", 6: "Suck 真空", 8: "Other 其他警示"}
ALCD_EN = {0: "Jam", 1: "Message", 4: "Cylinder", 5: "Motor",
           6: "Suck", 8: "Other"}


def load_alarms(csv_path):
    """AlarmList.csv (CP950, no BOM) -> list of dicts in S5F6 wire order.

    S5F6 order == std::map<AnsiString,...> order == plain lexicographic by code,
    which is exactly what uHGemHT160.cpp:3496 walks, so sorting here reproduces
    the wire order.
    """
    raw = open(csv_path, "rb").read()
    if raw[:3] == b"\xef\xbb\xbf":
        raise SystemExit("AlarmList.csv must not carry a UTF-8 BOM: " + csv_path)
    rows = [r for r in csv.reader(io.StringIO(raw.decode("cp950")))
            if r and r[0].strip()]
    if not rows or rows[0][:6] != CSV_COLS:
        raise SystemExit("unexpected AlarmList.csv header: %r"
                         % (rows[0][:6] if rows else None))
    out, seen = [], {}
    for r in rows[1:]:
        code = r[0].strip()
        if code in seen:
            raise SystemExit("duplicate AlarmCode in AlarmList.csv: " + code)
        seen[code] = 1
        cb = code.encode("cp950")
        alcd = int(r[1]) if r[1].strip() else 0
        cls, pay, why = classify(cb)
        out.append(dict(code=code, alcd=alcd, msg=r[2], cls=cls, payload=pay,
                        alid=cls * BAND + pay, why=why, old=legacy_hash(cb),
                        altx=(code + " " + r[2]) if r[2] else code,
                        label=ALCD_LABEL.get(alcd, "Other 其他警示")))
    out.sort(key=lambda d: d["code"])
    return out


def gate_csv(csv_path, alarms, expect_md5, expect_rows, allow_drift):
    """B1: refuse to derive anything from an AlarmList.csv that is not the one
    the constants (or the CLI overrides) vouch for."""
    got = hashlib.md5(open(csv_path, "rb").read()).hexdigest()
    problems = []
    if len(alarms) != expect_rows:
        problems.append("row count %d != expected %d" % (len(alarms), expect_rows))
    if got != expect_md5:
        problems.append("md5 %s != expected %s" % (got, expect_md5))
    if not problems:
        print("  csv gate : %d rows, md5 %s (verified)" % (len(alarms), got))
        return
    msg = ("AlarmList.csv gate FAILED (%s). This file is a boot artifact and may "
           "lag the source: boot the sim once, confirm the row count, then bump "
           "EXPECTED_CSV_ROWS / EXPECTED_CSV_MD5 in the same commit."
           % "; ".join(problems))
    if allow_drift:
        print("  !! WARNING (--allow-csv-drift): " + msg)
        return
    raise SystemExit("  !! " + msg)


# --------------------------------------------------------------------------- #
# AMENDMENT 2 -- offline twin of the startup self-check in database.cpp
# --------------------------------------------------------------------------- #
def audit(alarms, verbose=True):
    bad, byalid = [], {}
    for a in alarms:
        v, code = a["alid"], a["code"]
        if not (100000000 <= v <= 999999999):
            bad.append("%s -> %d is not exactly 9 digits" % (code, v))
        if v // BAND == 9:
            bad.append("%s -> %d is CLASS 9 (not decodable by the host); "
                       "register a numbered code for it" % (code, v))
        if v in byalid:
            bad.append("ALID %d collides: %s and %s" % (v, byalid[v], code))
        else:
            byalid[v] = code
        if decode(v)[2] != code:
            bad.append("%s -> %d does not round-trip (decodes to %r)"
                       % (code, v, decode(v)[2]))
    if verbose:
        vals = [a["alid"] for a in alarms]
        print("  rows              : %d" % len(alarms))
        print("  distinct ALIDs    : %d" % len(byalid))
        print("  class census      : %s"
              % dict(sorted(Counter(a["cls"] for a in alarms).items())))
        print("  ALCD census       : %s"
              % dict(sorted(Counter(a["alcd"] for a in alarms).items())))
        print("  min / max ALID    : %d / %d" % (min(vals), max(vals)))
        print("  all exactly 9 dig : %s" % all(len(str(v)) == 9 for v in vals))
        print("  round-trip exact  : %d/%d"
              % (sum(1 for a in alarms if decode(a["alid"])[2] == a["code"]),
                 len(alarms)))
        print("  above 2^31-1      : %d  (the signed-int32 warning is retired)"
              % sum(1 for v in vals if v > 2147483647))
        print("  changed vs legacy : %d" % sum(1 for a in alarms
                                               if a["alid"] != a["old"]))
        print("  ALTX > 40 chars   : %d  (separate issue, NOT fixed here; max %d)"
              % (sum(1 for a in alarms if len(a["altx"]) > 40),
                 max(len(a["altx"]) for a in alarms)))
        for c in sorted(CLASS_SLOTS):
            n = sum(1 for a in alarms if a["cls"] == c)
            print("  class %d %-22s : %3d in service / %d slots, headroom %d"
                  % (c, CLASS_NAME[c], n, CLASS_SLOTS[c], CLASS_SLOTS[c] - n))
        print("  violations        : %d" % len(bad))
    return bad


RATIFIED_SAMPLES = [
    ("40000", 400040000), ("50000", 500050000), ("60000", 600060000),
    ("70000", 700070000), ("80000", 800080000),
    ("MES1421", 300001421), ("MES0920", 300000920), ("JAM0913", 100000913),
    ("WAR0963", 200000963), ("WAR16120", 200016120), ("WAR16121", 200016121),
    ("SnFKCleanOut", 991410149),
    ("MES01421", 987714039), ("MES001421", 928189725), ("JAM913", 980936293),
    ("WAR09120", 952507070), ("MES0000", 300000000),
    ("7000", 901686169), ("700000", 920409945), ("90000", 954118329),
]


def selftest():
    """Every ratified sample plus the structural invariants, on strings alone."""
    bad = 0
    for code, want in RATIFIED_SAMPLES:
        got = optd(code.encode("cp950"))
        cls = got // BAND
        pay = got % BAND
        back = decode(got)[2]
        ok = got == want
        bad += 0 if ok else 1
        print("  %-14s -> %9d  class %d payload %-8d decode %-10s %s"
              % (code, got, cls, pay, back, "OK" if ok else "FAIL want %d" % want))
    n5 = ["%05d" % v for v in range(0, 100000)]
    w = [optd(s.encode()) for s in n5]
    if not all(len(str(x)) == 9 for x in w):
        print("  !! a five-digit numeric produced a non-9-digit ALID")
        bad += 1
    band48 = [x for s, x in zip(n5, w) if s[0] in "45678"]
    if len(band48) != 50000 or len(set(band48)) != 50000:
        print("  !! class 4..8 sweep not injective")
        bad += 1
    else:
        rt = all(decode(optd(s.encode()))[2] == s for s in n5 if s[0] in "45678")
        print("  class 4..8 sweep : 50,000 five-digit codes -> 50,000 distinct, "
              "all 9 digits, round-trip %s" % rt)
        bad += 0 if rt else 1
    can = []
    for p in ("JAM", "WAR", "MES"):
        can += ["%s%04d" % (p, v) for v in range(10000)]
        can += ["%s%05d" % (p, v) for v in range(10000, 100000)]
    cv = [optd(s.encode()) for s in can]
    if len(set(cv)) != 300000 or any(v // BAND == 9 for v in cv):
        print("  !! canonical class 1/2/3 sweep failed")
        bad += 1
    else:
        rt = all(decode(optd(s.encode()))[2] == s for s in can)
        print("  class 1/2/3 sweep: 300,000 canonical codes -> 300,000 distinct, "
              "0 demoted, round-trip %s" % rt)
        bad += 0 if rt else 1
    demote = ["MES01421", "MES001421", "JAM913", "WAR09120", "WAR016120",
              "MES14001421", "7000", "700000", "7000A", "90000", "30000"]
    if any(optd(s.encode()) // BAND != 9 for s in demote):
        print("  !! a string that must demote to class 9 did not")
        bad += 1
    else:
        print("  demotion set     : %d strings, all class 9" % len(demote))
    print("  selftest failures : %d" % bad)
    return bad


# --------------------------------------------------------------------------- #
# emitters
# --------------------------------------------------------------------------- #
SHEET_HEADER = ["ALID", "警報碼 AlarmCode", "ALCD", "類別", "目錄訊息 ALTX",
                "號段 Class", "號段內碼 Payload"]   # F/G layout, owner ruling 2026-09-03


def emit_sheet_csv(alarms, path):
    """(a) the workbook ALID-sheet rows, column order == the new 7-column sheet."""
    with io.open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, lineterminator="\r\n")
        w.writerow(SHEET_HEADER)
        for a in alarms:
            w.writerow([a["alid"], a["code"], a["alcd"], a["label"], a["altx"],
                        a["cls"], a["payload"]])
    return path


def emit_map_csv(alarms, path):
    """The review / delivery table (adds OldALID + Changed + decode round-trip)."""
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\r\n")
        w.writerow(["ALID", "Class", "Payload", "AlarmCode", "ALCD", "Category",
                    "ALTX", "OldALID", "Changed", "DecodesBackTo"])
        for a in alarms:
            w.writerow([a["alid"], a["cls"], a["payload"], a["code"], a["alcd"],
                        ALCD_EN.get(a["alcd"], "Other"), a["altx"], a["old"],
                        "YES" if a["alid"] != a["old"] else "NO",
                        decode(a["alid"])[2]])
    return path


SIM_BEGIN = "ALID_CATALOG = ["


def python_str(s):
    """ASCII-safe python literal; all 485 catalog ALTX values are ASCII."""
    if any(ord(c) > 126 for c in s):
        raise SystemExit("non-ASCII ALTX needs an explicit encoding: %r" % s)
    return '"%s"' % s.replace("\\", "\\\\").replace('"', '\\"')


def sim_block(alarms):
    """(b) the simulator ALID_CATALOG python block."""
    lines = [SIM_BEGIN]
    for a in alarms:
        lines.append("    (%d, %d, %s),"
                     % (a["alid"], a["alcd"], python_str(a["altx"])))
    lines.append("]")
    return "\n".join(lines) + "\n"


def emit_sim(alarms, sim_path, in_place):
    """Replace the ALID_CATALOG literal and every hard-coded row count."""
    src = io.open(sim_path, "r", encoding="utf-8", newline="").read()
    i = src.find("\n" + SIM_BEGIN)
    if i < 0:
        raise SystemExit("ALID_CATALOG block not found in " + sim_path)
    i += 1
    j = src.find("\n]\n", i)
    if j < 0:
        raise SystemExit("ALID_CATALOG terminator not found in " + sim_path)
    new = src[:i] + sim_block(alarms) + src[j + 3:]
    n = str(len(alarms))
    new = re.sub("目錄（[0-9]+ 筆）", "目錄（%s 筆）" % n, new)
    new = re.sub(r"(?<![0-9])481(?= 筆)", n, new)
    new = re.sub(r"the ([0-9]+)-entry catalog", "the %s-entry catalog" % n, new)
    new = re.sub(r"Alarm catalog, [0-9]+ rows", "Alarm catalog, %s rows" % n, new)
    if in_place:
        io.open(sim_path, "w", encoding="utf-8", newline="\n").write(new)
    return new


def emit_runner(alarms, runner_path, in_place):
    src = io.open(runner_path, "r", encoding="utf-8", newline="").read()
    new, k = re.subn(r"^ALARM_CATALOG_ROWS = [0-9]+",
                     "ALARM_CATALOG_ROWS = %d" % len(alarms), src, flags=re.M)
    if k != 1:
        raise SystemExit("ALARM_CATALOG_ROWS not found exactly once in "
                         + runner_path)
    if in_place:
        io.open(runner_path, "w", encoding="utf-8", newline="\n").write(new)
    return new


def emit_xlsx(alarms, base, out, expect_md5=None, prose=None):
    """Rebuild the ALID sheet of a COPY of `base`. Touches ONLY the cells it names.

    LAYOUT (owner ruling 2026-09-03, option F/G): columns A..E keep the 0831 order and
    meaning byte-for-byte (ALID | 警報碼 | ALCD | 類別 | 目錄訊息 ALTX); the two decode
    columns are APPENDED as F "號段 Class" and G "號段內碼 Payload". Nothing is inserted,
    so every host-side reference to "column B = alarm code" written against the 0831
    sheet stays valid.

    `prose` (optional dict, from --prose JSON):
        "alid_footnote_1" / "alid_footnote_2"  replace the two footnote paragraphs
        "cells": {"功能!E16": "...", ...}      overwrite arbitrary cells (value only,
                                               style untouched) - the customer-facing
                                               prose that the ALID re-encoding retires.

    HAZARDS handled here:
      * base md5 gate  -- the folder holds several same-shaped workbooks, plus a
        worktree copy tree under .claude/worktrees.
      * the two embedded PNGs live on the 功能 sheet (xl/media/image1.png,
        image2.png via xl/drawings/drawing1.xml). We insert no rows/cols anywhere.
        VERIFY the two media md5s after saving (--verify-png BASE OUT).
      * an open Excel lock file (~$...xlsx) means the base may still be dirty:
        close Excel before running this.
      * the footnote rows are FOUND, not assumed: last row whose column A is an int
        is the last data row; +1 blank spacer; +2 / +3 the two footnotes.
    """
    from copy import copy
    import openpyxl
    from openpyxl.styles import Alignment
    got = hashlib.md5(open(base, "rb").read()).hexdigest()
    if expect_md5 and got != expect_md5:
        raise SystemExit("base workbook md5 %s != expected %s -- wrong base file."
                         % (got, expect_md5))
    print("  base md5 %s (verified)" % got)
    wb = openpyxl.load_workbook(base)
    ws = wb["ALID"]
    last_data = max(r for r in range(2, ws.max_row + 1)
                    if isinstance(ws.cell(r, 1).value, int))
    OLD_F1, OLD_F2 = last_data + 2, last_data + 3
    foot1, foot2 = ws.cell(OLD_F1, 1).value, ws.cell(OLD_F2, 1).value
    if not (isinstance(foot1, str) and isinstance(foot2, str)):
        raise SystemExit("footnote rows not where expected (A%d / A%d)" % (OLD_F1, OLD_F2))
    print("  base ALID sheet: %d data rows, footnotes at A%d / A%d"
          % (last_data - 1, OLD_F1, OLD_F2))
    if prose:
        foot1 = prose.get("alid_footnote_1", foot1)
        foot2 = prose.get("alid_footnote_2", foot2)
    hdr_style = copy(ws.cell(1, 1)._style)
    body_style = [copy(ws.cell(2, c)._style) for c in range(1, 6)]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    for r in range(2, max(OLD_F2, len(alarms) + 6) + 1):
        for c in range(1, 8):
            ws.cell(r, c).value = None
    ws.cell(1, 6).value = SHEET_HEADER[5]
    ws.cell(1, 7).value = SHEET_HEADER[6]
    for c in (6, 7):
        ws.cell(1, c)._style = hdr_style
    style_for = {1: body_style[0], 2: body_style[1], 3: body_style[2],
                 4: body_style[3], 5: body_style[4], 6: body_style[0],
                 7: body_style[0]}
    for i, a in enumerate(alarms):
        r = i + 2
        vals = [a["alid"], a["code"], a["alcd"], a["label"], a["altx"],
                a["cls"], a["payload"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c)
            cell.value = v
            cell._style = style_for[c]
    f1 = len(alarms) + 3                 # blank spacer row at len+2
    f2 = f1 + 1
    ws.cell(f1, 1).value = foot1
    ws.cell(f2, 1).value = foot2
    for r in (f1, f2):
        ws.cell(r, 1)._style = copy(body_style[4])
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=f1, start_column=1, end_row=f1, end_column=7)
    ws.merge_cells(start_row=f2, start_column=1, end_row=f2, end_column=7)
    for col, wd in (("F", 10), ("G", 17)):   # A..E widths untouched
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A2"               # the 0831 file carried a stray A269
    ncell = 0
    if prose:
        for ref, text in (prose.get("cells") or {}).items():
            sheet, coord = ref.split("!", 1)
            wb[sheet][coord].value = text
            ncell += 1
    wb.save(out)
    print("  ALID sheet: %d data rows (2..%d); footnotes A%d / A%d; F/G decode "
          "columns; %d prose cell(s) written"
          % (len(alarms), len(alarms) + 1, f1, f2, ncell))
    return out


def verify_png(base, out):
    """The embedded-PNG hazard: prove openpyxl did not drop the two images."""
    import zipfile
    res = []
    for p in (base, out):
        z = zipfile.ZipFile(p)
        med = dict((n, hashlib.md5(z.read(n)).hexdigest())
                   for n in z.namelist() if n.startswith("xl/media/"))
        res.append(med)
    print("  base media : %s" % res[0])
    print("  out  media : %s" % res[1])
    ok = res[0] == res[1]
    print("  PNG round-trip byte-identical : %s" % ok)
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# prose guard -- claims that Option D retires
# --------------------------------------------------------------------------- #
RETIRED_PROSE = [
    (u"31[- ]?poly|31 進位|雜湊", "the 31-poly hash rule"),
    (u"無號|unsigned[- ]32|UNSIGNED 32", "the unsigned-32 parse warning"),
    (u"2\\^31|2147483647", "the signed-overflow warning"),
    (u"3184282107", "the old max ALID"),
    (u"2502907625|2054803979", "a retired hash example"),
    (u"3891410149", "a retired hash example (SnFKCleanOut is now 991410149)"),
    (u"4045923824", "a retired hash example (the live code is MES0920 = 300000920)"),
    (u"(?<![0-9])481(?![0-9])", "the stale 481 count"),
    (u"UsecegemMainFrom\\.cpp:1[3-6][0-9]-1[3-6][0-9]", "a stale source citation "
     "(ComputeAlarmAlid now sits at :167-177)"),
    (u"480\\+", "the vague '480+ codes' count"),
]


def check_prose(paths):
    hits = []
    for p in paths:
        if p.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(p)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for needle, what in RETIRED_PROSE:
                                if re.search(needle, cell.value):
                                    hits.append("%s  %s!%s asserts %s"
                                                % (os.path.basename(p), ws.title,
                                                   cell.coordinate, what))
        else:
            try:
                txt = io.open(p, encoding="utf-8").read()
            except (IOError, UnicodeDecodeError):
                continue
            for k, line in enumerate(txt.splitlines(), 1):
                for needle, what in RETIRED_PROSE:
                    if re.search(needle, line):
                        hits.append("%s:%d asserts %s"
                                    % (os.path.basename(p), k, what))
    return hits


def verify_table(alarms, path):
    """Cross-check against an externally produced old->new mapping CSV."""
    rows = list(csv.reader(io.open(path, encoding="utf-8")))
    hdr, body = rows[0], rows[1:]
    ic = dict((h, i) for i, h in enumerate(hdr))
    mine = dict((a["code"], a) for a in alarms)
    n = miss = diff = 0
    intable = set()
    for r in body:
        code = r[ic["AlarmCode"]]
        intable.add(code)
        if r[ic["Registered"]] != "YES":
            continue
        n += 1
        a = mine.get(code)
        if a is None:
            miss += 1
            print("  !! table row not in AlarmList.csv: %s" % code)
            continue
        if (int(r[ic["NewALID"]]) != a["alid"] or int(r[ic["Class"]]) != a["cls"]
                or int(r[ic["Payload"]]) != a["payload"]
                or int(r[ic["OldALID"]]) != a["old"]):
            diff += 1
            print("  !! %s table(%s cls %s pay %s old %s) != generator(%d %d %d %d)"
                  % (code, r[ic["NewALID"]], r[ic["Class"]], r[ic["Payload"]],
                     r[ic["OldALID"]], a["alid"], a["cls"], a["payload"], a["old"]))
    print("  registered rows compared : %d ; not in CSV : %d ; mismatches : %d"
          % (n, miss, diff))
    extra = set(mine) - intable
    print("  in AlarmList.csv but absent from the table : %d %s"
          % (len(extra), sorted(extra)[:5]))
    return 1 if (miss or diff or extra) else 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", default="D:\\HT160S_BCB")
    ap.add_argument("--csv")
    ap.add_argument("--expect-csv-md5", default=EXPECTED_CSV_MD5)
    ap.add_argument("--expect-csv-rows", type=int, default=EXPECTED_CSV_ROWS)
    ap.add_argument("--allow-csv-drift", action="store_true")
    ap.add_argument("--sim", default="D:\\AI_Area\\Tool\\HT160S_SECS_Simulator\\"
                                     "code\\secs_host_simulator.py")
    ap.add_argument("--runner", default="D:\\AI_Area\\Tool\\HT160S_SECS_Simulator\\"
                                        "code\\scenario_runner.py")
    ap.add_argument("--audit", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--emit-sheet-csv", metavar="OUT")
    ap.add_argument("--emit-map-csv", metavar="OUT")
    ap.add_argument("--emit-sim-block", metavar="OUT")
    ap.add_argument("--emit-sim", action="store_true")
    ap.add_argument("--emit-xlsx", metavar="OUT")
    ap.add_argument("--xlsx-base")
    ap.add_argument("--expect-base-md5")
    ap.add_argument("--prose", metavar="JSON",
                    help="UTF-8 JSON: alid_footnote_1 / alid_footnote_2 / cells{sheet!A1: text}")
    ap.add_argument("--verify-png", nargs=2, metavar=("BASE", "OUT"))
    ap.add_argument("--check-prose", nargs="+", metavar="FILE")
    ap.add_argument("--verify-table", metavar="CSV")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--check", action="store_true")
    a = ap.parse_args()

    if a.verify_png:
        sys.exit(verify_png(a.verify_png[0], a.verify_png[1]))
    csv_path = a.csv or os.path.join(a.repo, "system", "AlarmList.csv")
    alarms = load_alarms(csv_path)
    print("AlarmList.csv %s -> %d alarms" % (csv_path, len(alarms)))
    gate_csv(csv_path, alarms, a.expect_csv_md5, a.expect_csv_rows,
             a.allow_csv_drift)
    rc = 0
    doing = (a.emit_sheet_csv or a.emit_map_csv or a.emit_sim_block or a.emit_sim
             or a.emit_xlsx or a.check_prose or a.verify_table)
    if a.selftest:
        print("-- SELFTEST (ratified samples + structural sweeps) --")
        rc |= 1 if selftest() else 0
    if a.audit or a.check or not (doing or a.selftest):
        print("-- AMENDMENT 2 AUDIT (offline twin of the startup self-check) --")
        bad = audit(alarms)
        for b in bad:
            print("  !! " + b)
        rc |= 1 if bad else 0
    if a.emit_sheet_csv:
        print("wrote " + emit_sheet_csv(alarms, a.emit_sheet_csv))
    if a.emit_map_csv:
        print("wrote " + emit_map_csv(alarms, a.emit_map_csv))
    if a.emit_sim_block:
        io.open(a.emit_sim_block, "w", encoding="utf-8",
                newline="\n").write(sim_block(alarms))
        print("wrote %s (%d tuples)" % (a.emit_sim_block, len(alarms)))
    if a.emit_sim:
        emit_sim(alarms, a.sim, a.in_place)
        emit_runner(alarms, a.runner, a.in_place)
        print("simulator: %d ALID_CATALOG tuples + ALARM_CATALOG_ROWS=%d (%s)"
              % (len(alarms), len(alarms), "written" if a.in_place else "dry run"))
    if a.emit_xlsx:
        prose = None
        if a.prose:
            import json
            prose = json.loads(io.open(a.prose, encoding="utf-8-sig").read())
        print("wrote " + emit_xlsx(alarms, a.xlsx_base, a.emit_xlsx,
                                   a.expect_base_md5, prose))
    if a.check_prose:
        hits = check_prose(a.check_prose)
        print("-- PROSE GUARD --")
        for h in hits:
            print("  !! " + h)
        print("  %d retired claim(s) still present" % len(hits))
        rc |= 1 if hits else 0
    if a.verify_table:
        print("-- VERIFY against %s --" % a.verify_table)
        rc |= verify_table(alarms, a.verify_table)
    if a.check:
        for path, fn in ((a.sim, emit_sim), (a.runner, emit_runner)):
            cur = io.open(path, "r", encoding="utf-8", newline="").read()
            want = fn(alarms, path, False)
            ok = cur == want
            print("in sync: %-24s %s" % (os.path.basename(path), ok))
            rc |= 0 if ok else 1
    sys.exit(rc)


if __name__ == "__main__":
    main()

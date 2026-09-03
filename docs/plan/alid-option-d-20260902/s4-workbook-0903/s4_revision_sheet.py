# -*- coding: utf-8 -*-
# S4 step 2: apply the 修訂說明 (internal revision) sheet changes to an already-emitted workbook.
#   python s4_revision_sheet.py <workbook.xlsx> <s4_prose.json>
# Layout of the 0831/0903 sheet: R3 date, R4 file, R5 previous, R6 firmware, R7 "本次修訂內容",
# R8..R12 the 0831 items 1-5, R13 "尚待貴端確認 Open items". We insert the new items right
# after R7, then a history heading row, keep the old items verbatim, and append to open items.
import io, json, sys
from copy import copy
import openpyxl

wb_path, prose_path = sys.argv[1], sys.argv[2]
P = json.loads(io.open(prose_path, encoding="utf-8").read())
R = P["revision"]
wb = openpyxl.load_workbook(wb_path)
ws = wb["修訂說明"]
assert ws["A7"].value and "本次修訂內容" in str(ws["A7"].value), ws["A7"].value
assert ws["A13"].value and "尚待貴端確認" in str(ws["A13"].value), ws["A13"].value
assert str(ws["A8"].value).startswith("1. "), ws["A8"].value
for coord in ("B3", "B4", "B5", "B6"):
    ws[coord].value = R[coord]
a_style, b_style = copy(ws["A8"]._style), copy(ws["B8"]._style)
n_new = len(R["items"]) + 1          # items + history heading
ws.insert_rows(8, amount=n_new)
r = 8
for title, body in R["items"]:
    ws.cell(r, 1).value = title
    ws.cell(r, 2).value = body
    ws.cell(r, 1)._style = copy(a_style)
    ws.cell(r, 2)._style = copy(b_style)
    r += 1
ws.cell(r, 1).value = R["history_heading"]
ws.cell(r, 1)._style = copy(a_style)
ws.cell(r, 2).value = "（下列第 1～5 項為 2026-08-31 版的修訂內容，原文保留）"
ws.cell(r, 2)._style = copy(b_style)
open_row = 13 + n_new
assert "尚待貴端確認" in str(ws.cell(open_row, 1).value), ws.cell(open_row, 1).value
cur = ws.cell(open_row, 2).value or ""
ws.cell(open_row, 2).value = cur.rstrip("\n") + "\n" + R["open_item_append"]
wb.save(wb_path)
print("修訂說明: B3..B6 updated; %d new item rows + history heading inserted at R8..R%d; open items appended at R%d"
      % (len(R["items"]), 8 + n_new - 1, open_row))
